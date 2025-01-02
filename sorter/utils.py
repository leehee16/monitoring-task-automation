import os
import logging
from openpyxl import Workbook, load_workbook
from constants import IMAGE_EXTENSIONS

def setup_logging():
    logging.basicConfig(filename='image_classifier.log', level=logging.INFO,
                        format='%(asctime)s - %(levelname)s - %(message)s')

def load_excel_file(file_path):
    if os.path.exists(file_path):
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            classifications = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                user_id, classification_dates = row
                classification, *dates = classification_dates.split('_')
                classifications[user_id] = {
                    'classification': classification,
                    'problem_dates': dates[0].split(',') if dates else []
                }
            return classifications
        except Exception as e:
            logging.error(f"엑셀 파일 로드 중 오류 발생: {e}")
    return {}

def create_new_excel_file(file_path):
    try:
        wb = Workbook()
        ws = wb.active
        ws.append(['ID', 'Classification_Dates'])
        wb.save(file_path)
    except Exception as e:
        logging.error(f"새 엑셀 파일 생성 중 오류 발생: {e}")

def save_to_excel(file_path, classifications):
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        for user_id, data in classifications.items():
            classification = data['classification']
            problem_dates = data['problem_dates']
            dates_str = ','.join(problem_dates)
            classification_dates = f"{classification}_{dates_str}"
            ws.append([user_id, classification_dates])
        wb.save(file_path)
        logging.info(f"Excel file updated: {file_path}")
    except Exception as e:
        logging.error(f"엑셀 파일 저장 중 오류 발생: {e}")

def get_image_files(folder_path):
    return sorted([f for f in os.listdir(folder_path) if any(f.lower().endswith(ext) for ext in IMAGE_EXTENSIONS)])
