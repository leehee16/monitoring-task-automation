from PyQt5.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QScrollArea, QFileDialog, QInputDialog, QMessageBox, QProgressBar, QTabWidget, QTextEdit
from PyQt5.QtGui import QPixmap, QKeyEvent, QResizeEvent
from PyQt5.QtCore import Qt, QSize
from image_processor import ImageProcessor
from utils import load_excel_file, create_new_excel_file, save_to_excel, get_image_files
from constants import CLASSIFICATIONS, SETTINGS_FILE
import os
import json
import logging
from datetime import datetime
from pathlib import Path
import openpyxl

class ImageClassifier(QWidget):
    def __init__(self):
        super().__init__()
        self.initializeVariables()
        self.initUI()

    def initializeVariables(self):
        self.current_folder = None
        self.current_images = []
        self.current_index = 0
        self.classifications = {}
        self.user_folders = []
        self.current_user_index = 0
        self.problem_images = set()
        self.current_pixmap = None
        self.excel_file = None
        self.total_images = 0
        self.total_problem_images = 0
        
        # history 디렉토리 구조 설정
        try:
            # 현재 작업 디렉토리에서 history 디렉토리 찾기
            current_dir = Path(os.getcwd())
            while current_dir.name != 'history' and current_dir.parent != current_dir:
                current_dir = current_dir.parent
            
            if current_dir.name != 'history':
                raise ValueError("history 디렉토리를 찾을 수 없습니다.")
            
            self.history_dir = current_dir
            
            # 현재 주차 디렉토리 찾기 (YYYYMMDD-YYYYMMDD 형식)
            week_dirs = [d for d in self.history_dir.iterdir() 
                        if d.is_dir() and d.name[0].isdigit()]
            if week_dirs:
                self.week_dir = sorted(week_dirs)[-1]  # 가장 최근 주차
                
                # classified 디렉토리 설정
                self.classified_dir = self.week_dir / 'classified'
                self.classified_dir.mkdir(parents=True, exist_ok=True)
                
                # 엑셀 파일 경로 설정
                self.excel_file = self.week_dir / f"{self.week_dir.name}.xlsx"
            
        except Exception as e:
            logging.error(f"디렉토리 구조 초기화 중 오류 발생: {e}")

    def initUI(self):
        self.setWindowTitle('Image Classifier')
        self.setGeometry(100, 100, 1000, 800)

        # 메인 레이아웃
        main_layout = QVBoxLayout()

        # 탭 위젯 생성
        self.tab_widget = QTabWidget()

        # 메인 탭 생성
        main_tab = QWidget()
        main_tab_layout = QVBoxLayout()

        # 이미지 디스플레이 설정
        self.scroll_area = QScrollArea()
        self.image_label = QLabel()
        self.image_label.setAlignment(Qt.AlignCenter)
        self.scroll_area.setWidget(self.image_label)
        self.scroll_area.setWidgetResizable(True)
        self.scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        main_tab_layout.addWidget(self.scroll_area)

        # 상태 레이블 설정
        status_layout = QHBoxLayout()
        self.problem_label = QLabel('문제 있음: 아니오')
        self.progress_label = QLabel()
        self.completion_label = QLabel()
        self.history_label = QLabel('이전 분류: -')
        
        status_layout.addWidget(self.problem_label)
        status_layout.addWidget(self.progress_label)
        status_layout.addWidget(self.completion_label)
        status_layout.addWidget(self.history_label)
        main_tab_layout.addLayout(status_layout)

        # 버튼 설정
        self.select_folder_button = QPushButton('Select Folder')
        self.select_folder_button.clicked.connect(self.select_folder)
        main_tab_layout.addWidget(self.select_folder_button)

        self.save_button = QPushButton('Save to Excel')
        self.save_button.clicked.connect(self.save_to_excel)
        self.save_button.setEnabled(False)
        main_tab_layout.addWidget(self.save_button)

        # 프로그레스바 설정
        self.progress_bar = QProgressBar()
        main_tab_layout.addWidget(self.progress_bar)

        main_tab.setLayout(main_tab_layout)

        # 분석 탭 생성
        analysis_tab = AnalysisTab()

        # 탭 추가
        self.tab_widget.addTab(main_tab, "메인")
        self.tab_widget.addTab(analysis_tab, "분석")

        # 메인 레이아웃에 탭 위젯 추가
        main_layout.addWidget(self.tab_widget)

        # 메인 레이아웃 설정
        self.setLayout(main_layout)
        self.setFocusPolicy(Qt.StrongFocus)

    def load_settings(self):
        if os.path.exists(SETTINGS_FILE):
            try:
                with open(SETTINGS_FILE, 'r') as f:
                    settings = json.load(f)
                    last_folder = settings.get('last_folder')
                    if last_folder and os.path.exists(last_folder):
                        self.select_folder(last_folder)
            except Exception as e:
                logging.error(f"설정 파일 로드 중 오류 발생: {e}")

    def select_folder(self, folder=None):
        """폴더 선택 및 처리"""
        if not folder:
            folder = QFileDialog.getExistingDirectory(self, "Select Directory", self.current_folder)
            
        if folder:
            self.current_folder = folder
            # 폴더명을 기반으로 엑셀 파일명 생성
            folder_name = os.path.basename(folder)
            self.excel_file = os.path.join(folder, f"{folder_name}.xlsx")
            self.load_existing_classifications()
            self.load_user_folders()
            self.save_settings()
            self.start_image_processing()

    def load_existing_classifications(self):
        if os.path.exists(self.excel_file):
            self.classifications = load_excel_file(self.excel_file)
        else:
            create_new_excel_file(self.excel_file)

    def load_user_folders(self):
        self.user_folders = [f for f in os.listdir(self.current_folder) if os.path.isdir(os.path.join(self.current_folder, f))]
        self.user_folders = [f for f in self.user_folders if f not in self.classifications]
        self.current_user_index = 0
        self.total_images = 0
        self.total_problem_images = 0
        if self.user_folders:
            self.load_images()
        else:
            self.save_button.setEnabled(True)
            QMessageBox.information(self, '알림', '분류할 이미지가 없습니다.')

    def load_images(self):
        if self.current_user_index < len(self.user_folders):
            user_folder = self.user_folders[self.current_user_index]
            user_path = os.path.join(self.current_folder, user_folder)
            if os.path.exists(user_path):
                self.current_images = get_image_files(user_path)
                self.total_images += len(self.current_images)
                self.current_index = 0
                self.problem_images.clear()
                if self.current_images:
                    self.show_current_image()
                else:
                    self.delete_empty_folder(user_path)
                    self.next_user()
            else:
                # 폴더가 존재하지 않으면 다음 사용자로 이동
                self.next_user()
        else:
            self.save_button.setEnabled(True)
            QMessageBox.information(self, '완료', '모든 사용자 분류가 완료되었습니다.')

    def show_current_image(self):
        if self.current_images and self.current_index < len(self.current_images):
            user_folder = self.user_folders[self.current_user_index]
            image_path = os.path.join(self.current_folder, user_folder, self.current_images[self.current_index])
            self.current_pixmap = QPixmap(image_path)
            self.update_image_label()
            
            self.problem_label.setText(f"문제 있음: {'예' if self.current_images[self.current_index] in self.problem_images else '아니오'}")
            self.update_progress_label()
        else:
            self.image_label.clear()
            self.problem_label.setText("문제 있음: -")

    def update_image_label(self):
        if self.current_pixmap:
            scaled_pixmap = self.current_pixmap.scaled(
                self.scroll_area.size(),
                Qt.KeepAspectRatio,
                Qt.SmoothTransformation
            )
            self.image_label.setPixmap(scaled_pixmap)

    def update_progress_label(self):
        total = len(self.user_folders)
        current = self.current_user_index + 1
        self.progress_label.setText(f'진행 상황: {current}/{total}')

    def next_user(self):
        if not self.problem_images:
            reply = QMessageBox.question(self, '다음 폴더로 이동', 
                                         '문제가 있는 이미지가 없습니다. 다음 폴더로 이동하시겠습니까?',
                                         QMessageBox.Yes | QMessageBox.No, QMessageBox.Yes)
            if reply == QMessageBox.Yes:
                self.finalize_current_folder()
                self.move_to_next_folder()
            else:
                return
        else:
            self.classify_and_next_user()

    def classify_and_next_user(self):
        classification, ok = QInputDialog.getItem(self, "분류", "이 사용자의 분류를 선택하세요:", CLASSIFICATIONS, 0, False)
        if ok and classification:
            self.finalize_current_folder(classification)
            self.move_to_next_folder()
        else:
            return

    def finalize_current_folder(self, classification=None):
        if classification:
            # 원본 폴더 경로
            user_folder = self.user_folders[self.current_user_index]
            source_path = os.path.join(self.current_folder, user_folder)
            
            # classified 폴더 경로 (상위 디렉토리의 classified 폴더)
            history_week_dir = os.path.dirname(os.path.dirname(self.current_folder))
            classified_dir = os.path.join(history_week_dir, 'classified', user_folder)
            
            # classified 디렉토리 생성
            os.makedirs(classified_dir, exist_ok=True)
            
            # 문제가 있는 이미지만 classified 폴더로 이동
            for image in self.problem_images:
                source_file = os.path.join(source_path, image)
                target_file = os.path.join(classified_dir, image)
                try:
                    os.rename(source_file, target_file)
                    logging.info(f"이미지 이동: {source_file} -> {target_file}")
                except OSError as e:
                    logging.error(f"이미지 이동 실패: {e}")
            
            # 나머지 이미지 삭제
            self.delete_non_problem_images()
            
            # 분류 정보 저장
            self.save_classification(classification)
        else:
            self.delete_non_problem_images()

    def delete_non_problem_images(self):
        user_folder = self.user_folders[self.current_user_index]
        user_path = os.path.join(self.current_folder, user_folder)
        for image in self.current_images[:]:
            if image not in self.problem_images:
                image_path = os.path.join(user_path, image)
                try:
                    os.remove(image_path)
                    self.current_images.remove(image)
                    logging.info(f"이미지 삭제: {image_path}")
                except OSError as e:
                    logging.error(f"이미지를 삭제할 수 없습니다: {image_path}. 오류: {e}")
        
        if not self.current_images:
            self.delete_empty_folder(user_path)

    def delete_empty_folder(self, folder_path):
        try:
            os.rmdir(folder_path)
            logging.info(f"빈 폴더 삭제: {folder_path}")
        except OSError as e:
            logging.error(f"폴더를 삭제할 수 없습니다: {folder_path}. 오류: {e}")

    def move_to_next_folder(self):
        if self.current_user_index < len(self.user_folders) - 1:
            self.current_user_index += 1
            self.load_images()
        else:
            self.generate_report()
            QMessageBox.information(self, '완료', '모든 사용자 분류가 완료되었습니다.')
            self.save_button.setEnabled(True)

    def save_classification(self, classification):
        user_id = self.user_folders[self.current_user_index]
        problem_dates = sorted([img.split('_')[1].split('.')[0] for img in self.problem_images], reverse=True)
        self.classifications[user_id] = {
            'classification': classification,
            'problem_dates': problem_dates
        }
        
        # excel 파일 경로를 history 주간 폴더로 변경
        history_week_dir = os.path.dirname(os.path.dirname(self.current_folder))
        folder_name = os.path.basename(history_week_dir)
        self.excel_file = os.path.join(history_week_dir, f"{folder_name}.xlsx")
        
        save_to_excel(self.excel_file, {user_id: self.classifications[user_id]})

    def generate_report(self):
        # 리포트 파일 경로를 history 주간 폴더로 변경
        history_week_dir = os.path.dirname(os.path.dirname(self.current_folder))
        report_file = os.path.join(history_week_dir, f"classification_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
        
        try:
            with open(report_file, 'w') as f:
                f.write(f"작업 완료 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"총 이미지 수: {self.total_images}\n")
                f.write(f"문제 있는 이미지 수: {self.total_problem_images}\n")
                f.write(f"문제 비율: {self.total_problem_images / self.total_images * 100:.2f}%\n\n")
                
                for classification in CLASSIFICATIONS:
                    count = sum(1 for data in self.classifications.values() if data['classification'] == classification)
                    f.write(f"{classification} 분류 수: {count}\n")
            
            QMessageBox.information(self, '레포트 생성', f'레포트가 생성되었습니다: {report_file}')
        except Exception as e:
            logging.error(f"레포트 생성 중 오류 발생: {e}")
            QMessageBox.warning(self, '오류', '레포트를 생성하는 중 오류가 발생했습니다.')

    def save_settings(self):
        settings = {
            'last_folder': self.current_folder
        }
        try:
            with open(SETTINGS_FILE, 'w') as f:
                json.dump(settings, f)
        except Exception as e:
            logging.error(f"설정 파일 저장 중 오류 발생: {e}")

    def start_image_processing(self):
        self.image_processor = ImageProcessor(self.current_folder)
        self.image_processor.progress_updated.connect(self.update_progress_bar)
        self.image_processor.finished.connect(self.on_image_processing_finished)
        self.image_processor.start()

    def update_progress_bar(self, value):
        self.progress_bar.setValue(value)

    def on_image_processing_finished(self):
        QMessageBox.information(self, '처리 완료', '모든 이미지 처리가 완료되었습니다.')

    def keyPressEvent(self, event: QKeyEvent):
        if event.key() == Qt.Key_Left:  # 왼쪽 화살표 키
            if self.current_user_index > 0:
                prev_folder = self.user_folders[self.current_user_index - 1]
                prev_path = os.path.join(self.current_folder, prev_folder)
                
                if not os.path.exists(prev_path):
                    reply = QMessageBox.warning(
                        self,
                        '경고',
                        f'이전 폴더 ({prev_folder})가 이미 처리되어 삭제되었을 수 있습니다.\n계속 진행하시겠습니까?',
                        QMessageBox.Yes | QMessageBox.No,
                        QMessageBox.No
                    )
                    if reply == QMessageBox.No:
                        return
                
                self.current_user_index -= 1
                self.load_images()
                
        elif event.key() == Qt.Key_Right:  # 오른쪽 화살표 키
            if self.current_user_index < len(self.user_folders) - 1:
                self.current_user_index += 1
                self.load_images()
                
        elif event.key() == Qt.Key_Space:  # 스페이스바
            if self.current_images and self.current_index < len(self.current_images):
                current_image = self.current_images[self.current_index]
                if current_image in self.problem_images:
                    self.problem_images.remove(current_image)
                else:
                    self.problem_images.add(current_image)
                    self.total_problem_images += 1
                self.problem_label.setText(f"문제 있음: {'예' if current_image in self.problem_images else '아니오'}")
                
        event.accept()

    def resizeEvent(self, event: QResizeEvent):
        super().resizeEvent(event)
        if self.current_pixmap:
            self.show_current_image()

    def save_to_excel(self):
        if not self.classifications:
            QMessageBox.warning(self, '경고', '저장할 분류 데이터가 없습니다.')
            return

        try:
            save_to_excel(self.excel_file, self.classifications)
            QMessageBox.information(self, '저장 완료', f'분류 데이터가 {self.excel_file}에 저장되었습니다.')
        except Exception as e:
            logging.error(f"엑셀 파일 저장 중 오류 발생: {e}")
            QMessageBox.critical(self, '오류', '엑셀 파일 저장 중 오류가 발생했습니다.')

    def load_current_folder(self):
        """현재 선택된 폴더의 이미지들을 로드합니다."""
        if 0 <= self.current_folder_index < len(self.user_folders):
            self.current_folder = self.user_folders[self.current_folder_index]
            self.current_images = [f for f in self.current_folder.iterdir() 
                                 if f.suffix.lower() in ['.jpg', '.jpeg', '.png']]
            self.current_images.sort()
            
            if not self.current_images:  # 이미지가 없는 폴더는 다음으로 이동
                self.next_folder()
                return
                
            self.current_image_index = 0
            self.problem_images.clear()
            
            # 이전 분류 내역 확인
            user_id = self.current_folder.name
            history_text = '이전 분류: -'
            
            # history 디렉토리의 이전 주차 폴더들을 확인
            week_folders = sorted([d for d in self.history_dir.iterdir() 
                                 if d.is_dir() and d.name[0].isdigit()], reverse=True)
            
            previous_classifications = []
            for week_folder in week_folders:
                if week_folder == self.week_dir:  # 현재 주차는 건너뛰기
                    continue
                    
                excel_file = week_folder / f"{week_folder.name}.xlsx"
                if excel_file.exists():
                    try:
                        wb = openpyxl.load_workbook(excel_file)
                        ws = wb.active
                        
                        for row in ws.iter_rows(values_only=True):
                            if row and row[0] == user_id:
                                classification_data = row[1]
                                if '_' in classification_data:
                                    classification, dates = classification_data.split('_', 1)
                                else:
                                    classification = classification_data
                                    dates = ''
                                
                                week_start = week_folder.name.split('-')[0]
                                previous_classifications.append(f"{week_start}: {classification}")
                                break
                    except Exception as e:
                        self.logger.error(f"이전 분류 데이터 로드 중 오류 발생: {e}")
            
            if previous_classifications:
                history_text = '이전 분류: ' + ' | '.join(previous_classifications[-3:])  # 최근 3개만 표시
            
            self.history_label.setText(history_text)
            
            # 날짜 상태 초기화
            self.update_date_status()
            
            self.update_ui()
            self.show_current_image()
        else:
            self.finish_classification()

    def check_data_status(self):
        """데이터 상태를 확인합니다"""
        history_dir = Path(__file__).parent.parent.parent / 'history'
        self.current_week_dir = None
        self.data_status = {
            'has_history_folder': history_dir.exists(),
            'has_current_week': False,
            'has_data': False,
            'has_analysis': False,
            'has_classification': False
        }
        
        if self.data_status['has_history_folder']:
            week_folders = [d for d in history_dir.iterdir() if d.is_dir() and d.name[0].isdigit()]
            if week_folders:
                self.current_week_dir = sorted(week_folders)[-1]
                self.data_status['has_current_week'] = True
                
                # 데이터 존재 여부 확인
                data_dir = self.current_week_dir / 'data'
                self.data_status['has_data'] = data_dir.exists()
                
                # 분류 데이터 존재 여부 확인
                excel_file = self.current_week_dir / f"{self.current_week_dir.name}.xlsx"
                if excel_file.exists():
                    try:
                        wb = openpyxl.load_workbook(excel_file)
                        ws = wb.active
                        
                        # data 폴더의 모든 사용자 ID 가져오기
                        data_users = set()
                        if data_dir.exists():
                            data_users = {f.name for f in data_dir.iterdir() if f.is_dir()}
                        
                        # Excel에서 분류된 사용자 ID 가져오기
                        classified_users = set()
                        for row in ws.iter_rows(values_only=True):
                            if row and row[0]:  # 첫 번째 열이 사용자 ID
                                classified_users.add(row[0])
                        
                        # 모든 사용자가 분류되었는지 확인
                        self.data_status['has_classification'] = len(data_users - classified_users) == 0
                        
                    except Exception as e:
                        logging.error(f"분류 상태 확인 중 오류: {e}")

class AnalysisTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.initUI()
        
    def initUI(self):
        layout = QVBoxLayout()
        
        # 분석 버튼
        analyze_btn = QPushButton('분류 결과 분석')
        analyze_btn.clicked.connect(self.analyze_classifications)
        layout.addWidget(analyze_btn)
        
        # 결과 표시 영역
        self.result_text = QTextEdit()
        self.result_text.setReadOnly(True)
        layout.addWidget(self.result_text)
        
        self.setLayout(layout)
    
    def analyze_classifications(self):
        try:
            history_dir = Path(__file__).parent.parent.parent / 'history'
            week_folders = sorted([d for d in history_dir.iterdir() 
                                if d.is_dir() and d.name[0].isdigit()])
            
            analysis_results = []
            for week_dir in week_folders:
                excel_file = week_dir / f"{week_dir.name}.xlsx"
                if excel_file.exists():
                    wb = openpyxl.load_workbook(excel_file)
                    ws = wb.active
                    
                    week_stats = {
                        'week': week_dir.name,
                        'total_users': 0,
                        'classifications': {},
                        'problem_dates': set()
                    }
                    
                    for row in ws.iter_rows(values_only=True):
                        if row and len(row) >= 2:
                            week_stats['total_users'] += 1
                            classification_data = row[1]
                            
                            if '_' in classification_data:
                                classification, dates = classification_data.split('_', 1)
                                problem_dates = dates.split(',')
                                week_stats['problem_dates'].update(problem_dates)
                            else:
                                classification = classification_data
                            
                            week_stats['classifications'][classification] = \
                                week_stats['classifications'].get(classification, 0) + 1
                    
                    analysis_results.append(week_stats)
            
            # 결과 표시
            self.show_analysis_results(analysis_results)
            
        except Exception as e:
            QMessageBox.warning(self, '오류', f'분석 중 오류 발생: {str(e)}')
    
    def show_analysis_results(self, results):
        text = "=== 분류 결과 분석 ===\n\n"
        
        for week_stats in results:
            text += f"주차: {week_stats['week']}\n"
            text += f"총 사용자 수: {week_stats['total_users']}명\n"
            text += "분류 통계:\n"
            
            for classification, count in week_stats['classifications'].items():
                percentage = (count / week_stats['total_users']) * 100
                text += f"- {classification}: {count}명 ({percentage:.1f}%)\n"
            
            text += f"문제 발생 날짜 수: {len(week_stats['problem_dates'])}일\n"
            text += "\n"
        
        self.result_text.setText(text)

